"""
PROCAM GROUP — Performance Management System v5.0 (Production)
Flask + SQLite | 124 employees | 32 question sections | Procam branded
All data validated, EMP3482024 self-loop fixed, upload-ready

RUN: pip install flask flask-sqlalchemy flask-bcrypt flask-cors openpyxl gunicorn
     python app.py → http://localhost:5000
"""
import os, json, io, datetime
from flask import Flask, request, jsonify, session, send_file, render_template, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('DATABASE_PATH', os.path.join(BASE_DIR, 'procam_pms.db'))
app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'procam-pms-v5-prod-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(hours=8)
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
CORS(app, supports_credentials=True)


# ─── MODELS ──────────────────────────────────────────────────────────────────

class Company(db.Model):
    __tablename__ = 'companies'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    short_code = db.Column(db.String(20), unique=True, nullable=False)

class Location(db.Model):
    __tablename__ = 'locations'
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'))
    city = db.Column(db.String(80), nullable=False)
    state = db.Column(db.String(80))
    branch_name = db.Column(db.String(120))

class Grade(db.Model):
    __tablename__ = 'grades'
    id = db.Column(db.Integer, primary_key=True)
    grade_code = db.Column(db.String(10), unique=True, nullable=False)
    grade_name = db.Column(db.String(60), nullable=False)
    sort_order = db.Column(db.Integer, nullable=False)

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'))
    dept_name = db.Column(db.String(100), nullable=False)
    vertical_code = db.Column(db.String(20), nullable=False)

class Employee(db.Model):
    __tablename__ = 'employees'
    id = db.Column(db.Integer, primary_key=True)
    employee_code = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    designation = db.Column(db.String(100))
    grade_id = db.Column(db.Integer, db.ForeignKey('grades.id'))
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'))
    location_id = db.Column(db.Integer, db.ForeignKey('locations.id'))
    company_id = db.Column(db.Integer, db.ForeignKey('companies.id'))
    reporting_manager_id = db.Column(db.Integer, db.ForeignKey('employees.id'))
    date_of_joining = db.Column(db.String(20))
    email = db.Column(db.String(120))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    grade = db.relationship('Grade', foreign_keys=[grade_id])
    department = db.relationship('Department', foreign_keys=[department_id])
    location = db.relationship('Location', foreign_keys=[location_id])
    company = db.relationship('Company', foreign_keys=[company_id])
    manager = db.relationship('Employee', foreign_keys=[reporting_manager_id], remote_side='Employee.id')

class UserAuth(db.Model):
    __tablename__ = 'user_auth'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), unique=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default='EMPLOYEE')
    password_reset_required = db.Column(db.Boolean, default=True)
    failed_attempts = db.Column(db.Integer, default=0)
    last_login = db.Column(db.DateTime)
    employee = db.relationship('Employee')

class AppraisalCycle(db.Model):
    __tablename__ = 'appraisal_cycles'
    id = db.Column(db.Integer, primary_key=True)
    cycle_name = db.Column(db.String(60), nullable=False)
    fy_start = db.Column(db.String(20))
    fy_end = db.Column(db.String(20))
    parta_deadline = db.Column(db.String(20))
    partb_deadline = db.Column(db.String(20))
    results_publish_date = db.Column(db.String(20))
    status = db.Column(db.String(20), default='DRAFT')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class AppraisalForm(db.Model):
    __tablename__ = 'appraisal_forms'
    id = db.Column(db.Integer, primary_key=True)
    cycle_id = db.Column(db.Integer, db.ForeignKey('appraisal_cycles.id'))
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'))
    parta_status = db.Column(db.String(20), default='NOT_STARTED')
    parta_submitted_at = db.Column(db.DateTime)
    parta_last_saved = db.Column(db.DateTime)
    partb_status = db.Column(db.String(20), default='PENDING')
    partb_submitted_at = db.Column(db.DateTime)
    partb_last_saved = db.Column(db.DateTime)
    hr_status = db.Column(db.String(20), default='PENDING')
    final_score = db.Column(db.Float)
    final_rating = db.Column(db.String(40))
    increment_recommendation = db.Column(db.String(60))
    promotion_recommendation = db.Column(db.String(60))
    readiness_level = db.Column(db.String(60))
    hr_notes = db.Column(db.Text)
    employee = db.relationship('Employee')
    cycle = db.relationship('AppraisalCycle')

class PartARating(db.Model):
    __tablename__ = 'part_a_ratings'
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('appraisal_forms.id'))
    section_code = db.Column(db.String(10))
    question_index = db.Column(db.Integer)
    question_text = db.Column(db.Text)
    rating = db.Column(db.Integer)

class PartAText(db.Model):
    __tablename__ = 'part_a_texts'
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('appraisal_forms.id'))
    section_code = db.Column(db.String(10))
    field_key = db.Column(db.String(40))
    response_text = db.Column(db.Text)

class PartBRating(db.Model):
    __tablename__ = 'part_b_ratings'
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('appraisal_forms.id'))
    manager_id = db.Column(db.Integer)
    section_code = db.Column(db.String(10))
    question_index = db.Column(db.Integer)
    question_text = db.Column(db.Text)
    rating = db.Column(db.Integer)

class PartBText(db.Model):
    __tablename__ = 'part_b_texts'
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('appraisal_forms.id'))
    manager_id = db.Column(db.Integer)
    section_code = db.Column(db.String(10))
    field_key = db.Column(db.String(40))
    response_text = db.Column(db.Text)

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        return f(*args, **kwargs)
    return decorated


# ─── QUESTION BANK (32 sections: A1-A16, B1-B16) ────────────────────────────

QUESTION_BANK = {
    'A1': {'title': 'A1. Work Output & Goal Achievement', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': ['How well did you achieve the targets / KPIs set at the beginning of this appraisal period?','How consistently did you meet your deadlines and delivery commitments?','How would you rate the quality and accuracy of your work output?','How effectively did you manage multiple tasks or assignments simultaneously?','To what extent did you contribute to the overall objectives of your vertical / department?'],
        'open_text': [('a1_achievements','Describe 2–3 most significant achievements this year (with outcomes):'),('a1_gaps','Were there any targets not fully achieved? If so, state the reason:')]},
    'A2': {'title': 'A2. Behavioural & Professional Competencies', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': ['How would you rate your punctuality, attendance, and adherence to working hours?','How effectively did you communicate with team members, clients, and vendors?','How well did you collaborate and support colleagues across functions?','How would you rate your ability to adapt when priorities or situations changed?','How proactively did you identify problems and suggest or implement solutions?','How well did you handle pressure, stress, and difficult situations at work?','How effectively did you adhere to company policies, SOPs, and compliance requirements?','How would you rate your overall professional conduct and workplace behaviour?'],
        'open_text': [('a2_initiative','Describe a situation where you demonstrated initiative or went above and beyond:')]},
    'A3': {'title': 'A3. Learning & Development', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': ['How actively did you seek to learn new skills or knowledge relevant to your role?','How effectively did you apply new learnings to improve your work performance?','Did you participate in any training programs, certifications, or industry events this year?','How would you rate your readiness to take on additional responsibilities?'],
        'open_text': [('a3_training','List any training, certification, or skill development activity undertaken:'),('a3_goals','What skill or competency do you wish to develop in the coming year?')]},
    'A4': {'title': 'A4. Project Freight Management — PFM', 'subtitle': 'Function-specific — PFM vertical',
        'questions': ['How well did you meet your revenue / enquiry generation targets this year?','How effectively did you manage your existing client portfolio and ensure repeat business?','How would you rate your skills in prospecting, lead qualification, and deal closure?','How well did you coordinate with Operations to ensure smooth handover after deal closure?','How proactively did you gather and share market intelligence?','How effectively did you represent Procam capabilities and brand to clients?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a4_deals','State your key new accounts acquired or significant deals closed this year:')]},
    'A5': {'title': 'A5. Project Transport Management — PTM / PTM-M', 'subtitle': 'Function-specific — PTM vertical',
        'questions': ['How effectively did you plan and execute transport / project moves within agreed timelines?','How well did you coordinate with transporters, fleet, permits, and route planning teams?','How would you rate your documentation accuracy (permits, challans, trip sheets, eway bills)?','How proactively did you manage breakdowns, delays, and exceptions to minimise client impact?','How well did you maintain cost discipline and avoid unnecessary operational expenses?','How effectively did you use available tracking tools for fleet visibility and reporting?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a5_odc','Describe a complex ODC / project move you managed and what you did well:')]},
    'A6': {'title': 'A6. Engineering & Technical', 'subtitle': 'Function-specific — PTM + Installation',
        'questions': ['How effectively did you prepare route surveys, feasibility studies, or rigging plans?','How well did you ensure equipment selection was appropriate for cargo specifications?','How would you rate your adherence to safety standards and engineering best practices?','How proactively did you identify technical risks and communicate mitigation plans?','How well did you coordinate with operations and transport crews during critical moves?','How effectively did you maintain and update technical documentation and reports?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a6_challenge','Describe a technical challenge you resolved successfully this year:')]},
    'A7': {'title': 'A7. Warehouse & Logistics', 'subtitle': 'Function-specific — Warehouse vertical',
        'questions': ['How well did you maintain inventory accuracy and stock records?','How effectively did you manage inbound receipt, storage, and outbound despatch?','How well did you comply with safety, housekeeping, and 5S standards at the facility?','How proactively did you report damage, discrepancies, or safety hazards?','How efficiently did you utilise available storage space and handling equipment?','How well did you coordinate with operations, drivers, and client representatives?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a7_improvement','State any process improvement or initiative you implemented at the warehouse / yard:')]},
    'A8': {'title': 'A8. Corporate / Finance & Administration', 'subtitle': 'Function-specific — Corporate / Finance',
        'questions': ['How accurately and timely did you process invoices, payments, and reconciliations?','How effectively did you support the collections / outstanding recovery process?','How well did you comply with statutory requirements (GST, TDS, audit, etc.)?','How proactively did you flag financial irregularities or discrepancies?','How well did you support operational teams with cost data and analysis?','How would you rate your proficiency with accounting software and MIS reporting?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a8_improvement','State any significant process improvement or cost-saving initiative you contributed to:')]},
    'A9': {'title': 'A9. Driver & Transport Crew', 'subtitle': 'Function-specific — PTM Drivers/Operators/Helpers',
        'questions': ['How well did you adhere to assigned trip schedules and delivery timelines?','How would you rate your compliance with traffic rules, safety protocols, and load securing norms?','How effectively did you maintain your vehicle / equipment in good working condition?','How proactively did you report vehicle defects, route issues, or cargo damage?','How well did you coordinate with supervisors and operations team during trips?','How would you rate your conduct with clients, police officials, and highway authorities?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a9_incident','Describe any incident where you demonstrated responsible driving or handled a difficult situation:')]},
    'A10': {'title': 'A10. Rigging & Installation', 'subtitle': 'Function-specific — Installation vertical',
        'questions': ['How consistently did you meet the alignment standards without requirement of rework?','How good are you in adapting to sudden site challenges and unforeseen conditions?','How good are you in planning daily rigging operations and mobilising your team with clear instructions?','How good are you in Risk Assessment, Lifting Operations, Zero Incident Work, and Safety Documentation?','How well do you complete work within given timelines, client interfacing and coordination?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a10_challenge','Describe a technical challenge you resolved successfully this year:'),('a10_incident','Describe an incident during rigging / installation and how you handled it:'),('a10_client','Describe a client appreciation or commendation received this year:'),('a10_cost','Describe a cost optimisation exercise you implemented and its impact:')]},
    'A11': {'title': 'A11. Information Technology — IT', 'subtitle': 'Function-specific — IT',
        'questions': ['How effectively did you maintain IT infrastructure, networks, and systems with minimal downtime?','How promptly and satisfactorily did you resolve user support tickets and technical issues?','How well did you implement and maintain cybersecurity measures to protect company data and systems?','How proactively did you contribute to technology improvements, automation, and digital initiatives?','How effectively did you manage vendor relationships for software, hardware, and IT services?','How well did you maintain IT documentation, asset registers, and license compliance?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a11_achievements','Key IT achievements and projects completed or in progress this year:'),('a11_improvements','Describe any system improvement, automation, or digital initiative you led:')]},
    'A12': {'title': 'A12. Human Resources — HR', 'subtitle': 'Function-specific — HR',
        'questions': ['How effectively did you manage the recruitment and onboarding process end to end?','How well did you ensure compliance with labour laws, company HR policies, and statutory requirements?','How professionally did you handle employee relations, grievances, and disciplinary matters?','How well did you coordinate training and development programmes across the organisation?','How effectively did you maintain employee records, attendance systems, and HR documentation?','How proactively did you contribute to employee engagement, retention, and welfare initiatives?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a12_achievements','Key HR achievements this year:'),('a12_initiatives','HR initiatives planned or recommended for next year:')]},
    'A13': {'title': 'A13. Administration', 'subtitle': 'Function-specific — Admin + HR',
        'questions': ['How effectively did you manage office operations, facilities, and housekeeping?','How well did you coordinate travel, logistics, and accommodation arrangements for staff?','How proactively did you manage procurement of office supplies, stationery, and consumables?','How well did you maintain records, filing systems, and administrative documentation?','How effectively did you support management with coordination, scheduling, and correspondence?','How well did you manage vendor relationships for facility services (security, cleaning, maintenance)?','How well do you execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('a13_achievements','Key administrative achievements or improvements this year:')]},
    'A14': {'title': 'A14. HSE — Health, Safety & Environment', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': ['How well did you follow safety rules, wear prescribed PPE, and comply with site/office safety protocols?','How proactively did you report near-misses, unsafe conditions, or safety hazards at the workplace?','How effectively did you participate in safety training, toolbox talks, and emergency drills?','How would you rate your awareness and practice of first-aid, fire safety, and emergency procedures?','How well did you maintain housekeeping and cleanliness standards in your work area?'],
        'open_text': [('a14_incident','Describe any safety incident or near-miss you reported or managed this year:'),('a14_suggestion','Suggest any improvement to make the workplace safer or healthier:')]},
    'A15': {'title': 'A15. Sustainability & Environmental Responsibility', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': ['How well did you minimise waste (paper, fuel, electricity, water) in your daily work?','How effectively did you follow the company\'s waste segregation and disposal guidelines?','How proactively did you suggest or implement measures to reduce the environmental impact of operations?','How well did you ensure compliance with applicable environmental regulations and client sustainability requirements?'],
        'open_text': [('a15_initiative','Describe any sustainability or green initiative you contributed to this year:')]},
    'A16': {'title': 'A16. Goals for Next Year & Personal Development Plan', 'subtitle': 'Common — Mandatory for ALL employees',
        'questions': [],
        'open_text': [('a16_goals','State 3 professional goals you wish to achieve in the next appraisal year:'),('a16_support','What support, training, or resources do you need from the company to achieve these goals?'),('a16_feedback','Any other feedback or suggestions for improving the work environment at Procam:')]},

    'B1': {'title': 'B1. Work Output & Goal Achievement — Manager Assessment', 'subtitle': 'Manager rates employee',
        'questions': ['How well did the employee achieve the targets / KPIs set at the beginning of this appraisal period?','How consistently did the employee meet deadlines and delivery commitments?','How would you rate the quality and accuracy of the employee\'s work output?','How effectively did the employee manage multiple tasks or assignments simultaneously?','To what extent did the employee contribute to the overall objectives of the vertical / department?'],
        'open_text': [('b1_comment','Manager\'s comments on work output and goal achievement:')]},
    'B2': {'title': 'B2. Behavioural & Professional Competencies — Manager Assessment', 'subtitle': 'Manager rates employee',
        'questions': ['How would you rate the employee\'s punctuality, attendance, and adherence to working hours?','How effectively did the employee communicate with team members, clients, and vendors?','How well did the employee collaborate and support colleagues across functions?','How would you rate the employee\'s ability to adapt when priorities or situations changed?','How proactively did the employee identify problems and suggest or implement solutions?','How well did the employee handle pressure, stress, and difficult situations at work?','How effectively did the employee adhere to company policies, SOPs, and compliance requirements?','How would you rate the employee\'s overall professional conduct and workplace behaviour?'],
        'open_text': [('b2_comment','Manager\'s comments on behavioural and professional competencies:')]},
    'B3': {'title': 'B3. Learning & Development — Manager Assessment', 'subtitle': 'Manager rates employee',
        'questions': ['How actively did the employee seek to learn new skills or knowledge relevant to the role?','How effectively did the employee apply new learnings to improve work performance?','Did the employee participate in training programs, certifications, or industry events this year?','How would you rate the employee\'s readiness to take on additional responsibilities?'],
        'open_text': [('b3_comment','Manager\'s comments on learning and development:')]},
    'B4': {'title': 'B4. PFM — Manager Assessment', 'subtitle': 'Manager rates PFM employee',
        'questions': ['How well did the employee meet revenue / enquiry generation targets this year?','How effectively did the employee manage the client portfolio and ensure repeat business?','How would you rate the employee\'s skills in prospecting, lead qualification, and deal closure?','How well did the employee coordinate with Operations for smooth handover after deal closure?','How proactively did the employee gather and share market intelligence?','How effectively did the employee represent Procam capabilities and brand to clients?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b4_comment','Manager\'s assessment of PFM functional performance:')]},
    'B5': {'title': 'B5. PTM — Manager Assessment', 'subtitle': 'Manager rates PTM employee',
        'questions': ['How effectively did the employee plan and execute transport / project moves within agreed timelines?','How well did the employee coordinate with transporters, fleet, permits, and route planning teams?','How would you rate the employee\'s documentation accuracy (permits, challans, trip sheets, eway bills)?','How proactively did the employee manage breakdowns, delays, and exceptions to minimise client impact?','How well did the employee maintain cost discipline and avoid unnecessary operational expenses?','How effectively did the employee use available tracking tools for fleet visibility and reporting?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b5_comment','Manager\'s assessment of PTM functional performance:')]},
    'B6': {'title': 'B6. Engineering & Technical — Manager Assessment', 'subtitle': 'Manager rates Engineering employee',
        'questions': ['How effectively did the employee prepare route surveys, feasibility studies, or rigging plans?','How well did the employee ensure equipment selection was appropriate for cargo specifications?','How would you rate the employee\'s adherence to safety standards and engineering best practices?','How proactively did the employee identify technical risks and communicate mitigation plans?','How well did the employee coordinate with operations and transport crews during critical moves?','How effectively did the employee maintain and update technical documentation and reports?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b6_comment','Manager\'s assessment of Engineering functional performance:')]},
    'B7': {'title': 'B7. Warehouse — Manager Assessment', 'subtitle': 'Manager rates Warehouse employee',
        'questions': ['How well did the employee maintain inventory accuracy and stock records?','How effectively did the employee manage inbound receipt, storage, and outbound despatch?','How well did the employee comply with safety, housekeeping, and 5S standards at the facility?','How proactively did the employee report damage, discrepancies, or safety hazards?','How efficiently did the employee utilise available storage space and handling equipment?','How well did the employee coordinate with operations, drivers, and client representatives?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b7_comment','Manager\'s assessment of Warehouse functional performance:')]},
    'B8': {'title': 'B8. Corporate / Finance — Manager Assessment', 'subtitle': 'Manager rates Corporate/Finance employee',
        'questions': ['How accurately and timely did the employee process invoices, payments, and reconciliations?','How effectively did the employee support the collections / outstanding recovery process?','How well did the employee comply with statutory requirements (GST, TDS, audit, etc.)?','How proactively did the employee flag financial irregularities or discrepancies?','How well did the employee support operational teams with cost data and analysis?','How would you rate the employee\'s proficiency with accounting software and MIS reporting?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b8_comment','Manager\'s assessment of Corporate/Finance functional performance:')]},
    'B9': {'title': 'B9. Driver & Transport Crew — Manager Assessment', 'subtitle': 'Manager rates Driver/Crew employee',
        'questions': ['How well did the employee adhere to assigned trip schedules and delivery timelines?','How would you rate the employee\'s compliance with traffic rules, safety protocols, and load securing norms?','How effectively did the employee maintain the vehicle / equipment in good working condition?','How proactively did the employee report vehicle defects, route issues, or cargo damage?','How well did the employee coordinate with supervisors and operations team during trips?','How would you rate the employee\'s conduct with clients, police officials, and highway authorities?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b9_comment','Manager\'s assessment of Driver/Crew functional performance:')]},
    'B10': {'title': 'B10. Rigging & Installation — Manager Assessment', 'subtitle': 'Manager rates Installation employee',
        'questions': ['How consistently did the employee meet the alignment standards without requirement of rework?','How good is the employee in adapting to sudden site challenges and unforeseen conditions?','How good is the employee in planning daily rigging operations and mobilising the team with clear instructions?','How good is the employee in Risk Assessment, Lifting Operations, Zero Incident Work, and Safety Documentation?','How well does the employee complete work within given timelines, client interfacing and coordination?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b10_comment','Manager\'s assessment of Rigging & Installation functional performance:')]},
    'B11': {'title': 'B11. IT — Manager Assessment', 'subtitle': 'Manager rates IT employee',
        'questions': ['How effectively did the employee maintain IT infrastructure, networks, and systems with minimal downtime?','How promptly and satisfactorily did the employee resolve user support tickets and technical issues?','How well did the employee implement and maintain cybersecurity measures to protect company data and systems?','How proactively did the employee contribute to technology improvements, automation, and digital initiatives?','How effectively did the employee manage vendor relationships for software, hardware, and IT services?','How well did the employee maintain IT documentation, asset registers, and license compliance?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b11_comment','Manager\'s assessment of IT functional performance:')]},
    'B12': {'title': 'B12. HR — Manager Assessment', 'subtitle': 'Manager rates HR employee',
        'questions': ['How effectively did the employee manage the recruitment and onboarding process end to end?','How well did the employee ensure compliance with labour laws, company HR policies, and statutory requirements?','How professionally did the employee handle employee relations, grievances, and disciplinary matters?','How well did the employee coordinate training and development programmes across the organisation?','How effectively did the employee maintain employee records, attendance systems, and HR documentation?','How proactively did the employee contribute to employee engagement, retention, and welfare initiatives?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b12_comment','Manager\'s assessment of HR functional performance:')]},
    'B13': {'title': 'B13. Administration — Manager Assessment', 'subtitle': 'Manager rates Admin employee',
        'questions': ['How effectively did the employee manage office operations, facilities, and housekeeping?','How well did the employee coordinate travel, logistics, and accommodation arrangements for staff?','How proactively did the employee manage procurement of office supplies, stationery, and consumables?','How well did the employee maintain records, filing systems, and administrative documentation?','How effectively did the employee support management with coordination, scheduling, and correspondence?','How well did the employee manage vendor relationships for facility services (security, cleaning, maintenance)?','How well does the employee execute work within allocated budget using resources judiciously and controlling costs?'],
        'open_text': [('b13_comment','Manager\'s assessment of Admin functional performance:')]},
    'B14': {'title': 'B14. HSE — Manager Assessment', 'subtitle': 'Common — Manager rates employee on HSE',
        'questions': ['How well did the employee follow safety rules, wear prescribed PPE, and comply with site/office safety protocols?','How proactively did the employee report near-misses, unsafe conditions, or safety hazards at the workplace?','How effectively did the employee participate in safety training, toolbox talks, and emergency drills?','How would you rate the employee\'s awareness and practice of first-aid, fire safety, and emergency procedures?','How well did the employee maintain housekeeping and cleanliness standards in the work area?'],
        'open_text': [('b14_comment','Manager\'s comments on HSE performance:')]},
    'B15': {'title': 'B15. Sustainability — Manager Assessment', 'subtitle': 'Common — Manager rates employee on sustainability',
        'questions': ['How well did the employee minimise waste (paper, fuel, electricity, water) in daily work?','How effectively did the employee follow the company\'s waste segregation and disposal guidelines?','How proactively did the employee suggest or implement measures to reduce the environmental impact of operations?','How well did the employee ensure compliance with applicable environmental regulations and client sustainability requirements?'],
        'open_text': [('b15_comment','Manager\'s comments on sustainability performance:')]},
    'B16': {'title': 'B16. Overall Assessment & Recommendation', 'subtitle': 'Manager provides overall assessment',
        'questions': [],
        'open_text': [('b16_strengths','Key strengths of this employee:'),('b16_development','Areas where improvement is needed:'),('b16_readiness','Readiness for promotion / additional responsibility (with rationale):'),('b16_recommendation','Overall recommendation (increment / promotion / role change / training):'),('b16_comments','Any other comments or observations:')]},
}


# ─── EMPLOYEE MASTER DATA (124 validated employees) ────────────────────────

# Format: (code, name, vertical, sub_function, grade, designation, manager_code, role, part_a_sections, part_b_sections)

EMPLOYEE_MASTER = [

    ('HR001', 'HR Admin', 'Corporate', 'HR', 'M1', 'HR Administrator', '', 'HR_ADMIN', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('DIR12010', 'NILESH KUMAR SINHA', 'Corporate', 'Corporate', 'M1', 'Director / SUPER ADMIN', 'HR001', 'SUPER_ADMIN', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('DIR22010', 'JAMES FRANCIS XAVIER', 'PTM', 'PTM', 'M1', 'Director / HR ADMIN', 'DIR12010', 'HR_ADMIN', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('DIR42010', 'T G RAMALINGAM', 'Corporate', 'Finance', 'M1', 'Director / HR ADMIN', 'HR001', 'HR_ADMIN', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('DIR52011', 'SETHUPATHY SUNDARAM', 'Installation', 'Installation', 'M1', 'Director / HR ADMIN', 'DIR12010', 'HR_ADMIN', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('DIR72012', 'SRINIVAS MARELLA', 'Warehouse', 'Warehouse', 'M1', 'Director / HR ADMIN', 'DIR12010', 'HR_ADMIN', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('CON252022', 'SAMSUL HUSSAIN', 'Installation', 'Installation', 'C2', 'Project Consultant', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP472012', 'GOWDHAMAN RAJAKRISHNAN', 'Installation', 'Installation', 'M2', 'General Manager', 'DIR52011', 'MANAGER', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP1282017', 'PRAVINKUMAR ARUMUGAM', 'Installation', 'Installation', 'M3', 'Manager', 'DIR52011', 'MANAGER', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP1332017', 'SANTOSH KUMAR', 'Installation', 'Installation', 'E1', 'Executive', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP2832022', 'MOHANRAJ R', 'Installation', 'Installation', 'E1', 'Asst Manager', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3282023', 'JONES GEORGE T', 'Installation', 'Installation', 'T1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3292023', 'ZAHID KHAN', 'Installation', 'Installation', 'M3', 'Project Manager', 'EMP472012', 'MANAGER', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3382024', 'YOGESH KUMAR RAJASEKARAN', 'Installation', 'Installation', 'T1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3552024', 'KARTHIKEYAN  R', 'Installation', 'Installation', 'T1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3602024', 'AHMAD ALI', 'Installation', 'Installation', 'J1', 'Supervisor', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3612024', 'KAMAR KHAN', 'Installation', 'Installation', 'T1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3672025', 'AKASH PRABU', 'Installation', 'Installation', 'J1', 'HSE Officer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3742025', 'MUNTAZIR ALAM', 'Installation', 'Installation', 'J1', 'HSE Officer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3752025', 'VIKASH DUBEY', 'Installation', 'Installation', 'E1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3882025', 'SUNDHAR RAJAN S', 'Installation', 'Installation', 'E1', 'Project Engineer', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP3912025', 'SHYAM BHARTI', 'Installation', 'Installation', 'J2', 'Supervisor', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP4002025', 'MD  INAMUDDIN', 'Installation', 'Installation', 'J1', 'HSE Officer', 'EMP3292023', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP4032025', 'PRAMOD KUMAR', 'Installation', 'Installation', 'J1', 'Sr. Supervisor', 'EMP3292023', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP4042026', 'GURUSWAMI MOHANTA', 'Installation', 'Installation', 'J1', 'Supervisor', 'EMP1282017', 'EMPLOYEE', 'A1, A2, A3, A6, A10, A14, A15, A16', 'B1, B2, B3, B6, B10, B14, B15, B16'),

    ('EMP12010', 'NITIN RAWAT', 'PTM', 'PTM', 'M1', 'Asst Vice President', 'DIR12010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('CON242017', 'LAKSHMI NARAYAN', 'PTM', 'PTM', 'C2', 'Consultant', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP132010', 'SAHADEB SAHOO', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP142010', 'RANJIT GOGOI', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP162010', 'MOHD RAHIMUDDIN RAHIMUDDIN', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP172010', 'MANJURUL HOQUE', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP212010', 'AJIT KUMAR DAS', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP572012', 'VIJAY T V', 'PTM', 'PTM', 'M2', 'Dy. General Manager', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP812015', 'RAMESH YADAV SECHAE', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1062016', 'KAMRUL ISLAM', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1082016', 'NISHIT RANJAN DAS', 'PTM', 'PTM', 'J1', 'Operator', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1112016', 'BIJOY KONWAR', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1172016', 'NARENDER SINGH', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1342017', 'PRAMOD KUMAR SUKLA', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1372017', 'DINESH DUBEY', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1472018', 'KULDIP KUMAR', 'PTM', 'PTM', 'J1', 'Driver', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1482018', 'PRATAP SINGH', 'PTM', 'PTM', 'J1', 'Driver -Blade', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1492018', 'CHAKRADHAR SAHOO', 'PTM', 'PTM', 'J2', 'Assistant', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1502018', 'DEVENDRA SUBHASH', 'PTM', 'PTM', 'J1', 'Driver -Blade', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP1552018', 'ABHISHEK SINGH', 'PTM', 'PTM', 'E1', 'Executive', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1612018', 'AMIT KUMAR', 'PTM', 'PTM', 'J2', 'Assistant', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1662018', 'PHOOL CHANDRA YUDHISHIR', 'PTM', 'PTM', 'J1', 'Assistant', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1672018', 'PARTAB SINGH', 'PTM', 'PTM', 'J1', 'Assistant', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1682018', 'TAHIRUL HAQUE', 'PTM', 'PTM', 'J1', 'Assistant', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1702018', 'SURENDAR SINGH', 'PTM', 'PTM', 'J1', 'Driver -Blade', 'EMP3482024', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A9, A14, A15, A16', 'B1, B2, B3, B5, B6, B9, B14, B15, B16'),

    ('EMP2122018', 'GAJENDRA KUMAR GIRI', 'PTM', 'PTM', 'J1', 'Supervisor', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3022023', 'BHUSHAN B BHAGAT', 'PTM', 'PTM', 'M3', 'Manager', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3482024', 'SHRIRAM DATTU PATIL', 'PTM', 'PTM', 'M3', 'Manager', 'DIR22010', 'MANAGER', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3902025', 'BALA MURUGAN T', 'PTM', 'PTM', 'M2', 'Manager', 'DIR22010', 'MANAGER', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3932025', 'MANISH KUMAR BHAKTA', 'PTM', 'PTM', 'T1', 'HSE Officer', 'EMP3902025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3952025', 'VENKATESH RAMARAO ALTHADA', 'PTM', 'PTM', 'E1', 'Manager', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3982025', 'ASHITOSH SARJERAO GHOLAP', 'PTM', 'PTM', 'J2', 'Supervisor', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3992025', 'HAZARAT ALI', 'PTM', 'PTM', 'J1', 'Supervisor', 'EMP3902025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP22010', 'RAJEEV RANJAN', 'PTM-M', 'PTM-M', 'E1', 'Executive', 'EMP3702025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP1322017', 'RAM MOHAN CHAUBEY', 'PTM-M', 'PTM-M', 'E1', 'Executive', 'DIR22010', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3142023', 'BALKRISHNAN SHARMA', 'PTM-M', 'PTM-M', 'J1', 'Supervisor', 'EMP3702025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3532024', 'SURESH  KUMAR', 'PTM-M', 'PTM-M', 'J1', 'Executive', 'EMP3702025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3702025', 'SURANJAN AON', 'PTM-M', 'PTM-M', 'M2', 'Dy. General Manager', 'DIR22010', 'MANAGER', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('EMP3862025', 'CHANDRESH KUMAR BAIJNATH YADAV', 'PTM-M', 'PTM-M', 'J2', 'Operator', 'EMP3702025', 'EMPLOYEE', 'A1, A2, A3, A5, A6, A14, A15, A16', 'B1, B2, B3, B5, B6, B14, B15, B16'),

    ('CON142024', 'SEEMA SANJEEV MOGHE', 'PFM', 'PFM', 'C2', 'Consultant - Full Time', 'EMP192010', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP112010', 'SANJEEV KUMAR PALIWAL', 'PFM', 'PFM', 'M2', 'General Manager', 'DIR12010', 'MANAGER', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP192010', 'MANJU MISHRA', 'PFM', 'PFM', 'M2', 'Sr Manager', 'DIR12010', 'MANAGER', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP242010', 'LAXMI RAM SINGH', 'PFM', 'PFM', 'M3', 'Manager', 'EMP372011', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP372011', 'SANJNA VARDHAN', 'PFM', 'PFM', 'M2', 'General Manager', 'DIR12010', 'MANAGER', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP1602018', 'SANTHOSH P', 'PFM', 'PFM', 'J2', 'Assistant', 'DIR52011', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP2582020', 'DATTARAM MAHALIM', 'PFM', 'PFM', 'E1', 'Executive', 'EMP192010', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP2642021', 'KUMAR SATYAM RAY', 'PFM', 'PFM', 'E1', 'Executive', 'EMP112010', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP2752021', 'SAGAR BHOGLE', 'PFM', 'PFM', 'E1', 'Executive', 'EMP112010', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP2982023', 'SHARAYU UDAY BHOSALE', 'PFM', 'PFM', 'E1', 'Executive', 'EMP372011', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP3212023', 'SACHIN THAKUR', 'PFM', 'PFM', 'E1', 'Customer Service Executive', 'EMP112010', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP3322023', 'ARYAAN  SHAIKH', 'PFM', 'PFM', 'E1', 'Executive', 'EMP372011', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('EMP3592024', 'AMIT KAKKAR', 'PFM', 'PFM', 'M3', 'Manager', 'EMP372011', 'EMPLOYEE', 'A1, A2, A3, A4, A14, A15, A16', 'B1, B2, B3, B4, B14, B15, B16'),

    ('CON1362025', 'PRAVIN  CHOUDHARY', 'Warehouse', 'Warehouse', 'M2', 'Consultant - Full Time', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2782022', 'GAJANAN NARAYAN NAGLOT', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2792022', 'SWAPNIL SUNIL JADHAV', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2802022', 'AMOL BHAGVAN NIKAM', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2892022', 'RAKESH DNYANESHWAR RAWAL', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2902022', 'VIPUL SINH ZALA', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2952023', 'RAMESHWAR NIHALSINGH GUSINGE', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP2962023', 'SHASHIDHAR PANDURANG NAIK', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3042023', 'VISHAL RAOSAHEB MAGAR', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3062023', 'NITIN AMBADAS PAWAR', 'Warehouse', 'Warehouse', 'J1', 'Sr. Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3072023', 'SOHEL MAINOOR SHAIKH', 'Warehouse', 'Warehouse', 'J1', 'HSE Officer', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3092023', 'SATISH DATTA NAVGHARE', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3102023', 'BALU BHAGOVRAO JOGDANAD', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3152023', 'ANURAG UDAY CHAND', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3162023', 'BIRENDRA KUMAR', 'Warehouse', 'Warehouse', 'E1', 'Asst Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3192023', 'PANJAB DINKAR PISE', 'Warehouse', 'Warehouse', 'J1', 'Data Entry Operator', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3202023', 'SHIVAJI ASHOK DHUMAL', 'Warehouse', 'Warehouse', 'J1', 'Data Entry Operator', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3372024', 'SAYANTAN NASKAR', 'Warehouse', 'Warehouse', 'T1', 'SITE ENGINEER', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3642024', 'SOUVIK CHAKRABORTY', 'Warehouse', 'Warehouse', 'J1', 'HSE Officer', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3762025', 'PARVEEN SHARMA', 'Warehouse', 'Warehouse', 'J1', 'Sr. Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3772025', 'SANJAY BHITE', 'Warehouse', 'Warehouse', 'M3', 'Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3782025', 'SATISH JADHAV', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3802025', 'KAPIL BEKANALE', 'Warehouse', 'Warehouse', 'J1', 'Sr. Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3822025', 'AKASH SOMNATH NARAYNE', 'Warehouse', 'Warehouse', 'J1', 'Sr. Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3832025', 'VISHAL PUNDLIK BHOKRE', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3842025', 'SAURABH RAMESH WAGHMARE', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP3852025', 'AVINASH TUKARAM GHATUL', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP4022025', 'VIKRANT VATS', 'Warehouse', 'Warehouse', 'J2', 'Sr. Manager', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP4052026', 'AKRAM MAHMUD MUJAWAR', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP4062026', 'PRAVIN ABASAHEB BARDE', 'Warehouse', 'Warehouse', 'J1', 'Supervisor', 'DIR72012', 'EMPLOYEE', 'A1, A2, A3, A7, A14, A15, A16', 'B1, B2, B3, B7, B14, B15, B16'),

    ('EMP182010', 'K UMAMAHESWARA RAO', 'Corporate', 'Finance', 'M2', 'Dy. General Manager', 'DIR42010', 'MANAGER', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP482012', 'SAKTHEESWARI MURUGAVEL', 'Corporate', 'IT', 'M3', 'Sr Manager', 'DIR12010', 'EMPLOYEE', 'A1, A2, A3, A11, A14, A15, A16', 'B1, B2, B3, B11, B14, B15, B16'),

    ('EMP2482019', 'TANIMA MUKHERJEE', 'Corporate', 'HR', 'M3', 'Sr Manager', 'DIR42010', 'MANAGER', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('EMP2622020', 'ARITRA MITRA', 'Corporate', 'Finance', 'J1', 'Supervisor', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP2882022', 'SEEMA CHATTOPADHYAY', 'Corporate', 'Admin', 'E1', 'Executive Business Administrator', 'DIR12010', 'EMPLOYEE', 'A1, A2, A3, A13, A14, A15, A16', 'B1, B2, B3, B13, B14, B15, B16'),

    ('EMP2972023', 'JAYANTA KUMAR PAUL', 'Corporate', 'HR', 'E1', 'Executive', 'EMP2482019', 'EMPLOYEE', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('EMP2992023', 'DIPANKA TALUKDER', 'Corporate', 'Admin', 'E1', 'Executive', 'DIR12010', 'EMPLOYEE', 'A1, A2, A3, A13, A14, A15, A16', 'B1, B2, B3, B13, B14, B15, B16'),

    ('EMP3122023', 'SUNITA NAGA ALKAR', 'Corporate', 'Finance', 'J2', 'Assistant', 'EMP182010', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3132023', 'BIDISHA BANERJEE', 'Corporate', 'Finance', 'J1', 'Accounts Supervisor', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3222023', 'SAYANTI  GHOSH', 'Corporate', 'Finance', 'J1', 'Accounts Supervisor', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3542024', 'SAYAN  DAS', 'Corporate', 'HR', 'E1', 'Executive', 'EMP2482019', 'EMPLOYEE', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('EMP3652025', 'SUMIT MONDAL', 'Corporate', 'Finance', 'J1', 'Accountant', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3732025', 'BIKASH  ROUTH', 'Corporate', 'HR', 'J2', 'Assistant', 'EMP2482019', 'EMPLOYEE', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('EMP3892025', 'BHAVIN VINODHBHAI JIILKA', 'Corporate', 'Finance', 'M1', 'Head of Accounts & Finance', 'DIR42010', 'MANAGER', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3942025', 'ANIKET RAY CHAUDHURI', 'Corporate', 'HR', 'E1', 'Executive', 'EMP2482019', 'EMPLOYEE', 'A1, A2, A3, A12, A13, A14, A15, A16', 'B1, B2, B3, B12, B13, B14, B15, B16'),

    ('EMP3962025', 'DHANASHREE HARISHCHANDRA PAWAR', 'Corporate', 'Finance', 'E1', 'Accountant', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

    ('EMP3972025', 'SAMIKSHA CHANDRAKANT VAYNGANKAR', 'Corporate', 'Finance', 'J2', 'Accounts Supervisor', 'EMP3892025', 'EMPLOYEE', 'A1, A2, A3, A8, A14, A15, A16', 'B1, B2, B3, B8, B14, B15, B16'),

]


# Build lookup dict
EMP_LOOKUP = {}
for _e in EMPLOYEE_MASTER:
    EMP_LOOKUP[_e[0]] = {
        'n':_e[1], 'v':_e[2], 'sf':_e[3], 'g':_e[4], 'd':_e[5], 'mc':_e[6], 'r':_e[7],
        'a_secs': [s.strip() for s in _e[8].split(',')],
        'b_secs': [s.strip() for s in _e[9].split(',')],
    }

MANAGERS_WITH_REPORTS = set()
for _e in EMPLOYEE_MASTER:
    if _e[6]:
        MANAGERS_WITH_REPORTS.add(_e[6])


# ─── AUTH ROUTES ─────────────────────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json or {}
    code = (data.get('employee_code') or '').strip().upper()
    pw = data.get('password') or ''
    if not code or not pw:
        return jsonify({'error': 'Please enter employee code and password'}), 401
    emp = Employee.query.filter_by(employee_code=code, is_active=True).first()
    if not emp:
        return jsonify({'error': 'Invalid employee code or account inactive'}), 401
    ua = UserAuth.query.filter_by(employee_id=emp.id).first()
    if not ua:
        return jsonify({'error': 'Account not set up. Contact HR.'}), 401
    if ua.failed_attempts and ua.failed_attempts >= 5:
        return jsonify({'error': 'Account locked. Contact HR.'}), 401
    if not bcrypt.check_password_hash(ua.password_hash, pw):
        ua.failed_attempts = (ua.failed_attempts or 0) + 1
        db.session.commit()
        return jsonify({'error': 'Incorrect password'}), 401
    ua.failed_attempts = 0
    ua.last_login = datetime.datetime.utcnow()
    db.session.commit()
    session.permanent = True
    session['user_id'] = ua.id
    session['employee_id'] = emp.id
    session['role'] = ua.role
    session['employee_code'] = emp.employee_code
    md = EMP_LOOKUP.get(code, {})
    return jsonify({
        'success': True, 'employee_code': emp.employee_code, 'full_name': emp.full_name,
        'role': ua.role, 'password_reset_required': ua.password_reset_required,
        'vertical_code': md.get('v', ''), 'designation': md.get('d', emp.designation),
        'grade': md.get('g', ''), 'sub_function': md.get('sf', ''),
        'manager_code': md.get('mc', ''), 'manager_name': EMP_LOOKUP.get(md.get('mc',''),{}).get('n',''),
        'question_sections_a': md.get('a_secs', []),
        'question_sections_b': md.get('b_secs', []),
    })

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/me', methods=['GET'])
def me():
    if 'user_id' not in session:
        return jsonify({'authenticated': False})
    ua = UserAuth.query.get(session['user_id'])
    if not ua:
        return jsonify({'authenticated': False})
    emp = ua.employee
    md = EMP_LOOKUP.get(emp.employee_code, {})
    return jsonify({
        'authenticated': True, 'employee_code': emp.employee_code, 'full_name': emp.full_name,
        'role': ua.role, 'password_reset_required': ua.password_reset_required,
        'vertical_code': md.get('v', ''), 'designation': md.get('d', emp.designation),
        'grade': md.get('g', ''), 'sub_function': md.get('sf', ''),
        'manager_code': md.get('mc', ''), 'manager_name': EMP_LOOKUP.get(md.get('mc',''),{}).get('n',''),
        'question_sections_a': md.get('a_secs', []),
        'question_sections_b': md.get('b_secs', []),
        'employee_id': emp.id,
    })

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json or {}
    ua = UserAuth.query.get(session['user_id'])
    if not ua.password_reset_required:
        if not bcrypt.check_password_hash(ua.password_hash, data.get('current_password', '')):
            return jsonify({'error': 'Current password incorrect'}), 400
    pw = data.get('new_password', '')
    if len(pw) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    ua.password_hash = bcrypt.generate_password_hash(pw).decode()
    ua.password_reset_required = False
    db.session.commit()
    return jsonify({'success': True})


# ─── DASHBOARD ───────────────────────────────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard():
    active = AppraisalCycle.query.filter_by(status='ACTIVE').first()
    if not active: return jsonify({'cycle': None, 'stats': {}})
    forms = AppraisalForm.query.filter_by(cycle_id=active.id).all()
    total = len(forms)
    parta_sub = sum(1 for f in forms if f.parta_status == 'SUBMITTED')
    parta_draft = sum(1 for f in forms if f.parta_status == 'IN_PROGRESS')
    parta_not = sum(1 for f in forms if f.parta_status == 'NOT_STARTED')
    partb_sub = sum(1 for f in forms if f.partb_status == 'SUBMITTED')
    partb_draft = sum(1 for f in forms if f.partb_status == 'IN_PROGRESS')
    hr_scored = sum(1 for f in forms if f.hr_status == 'SCORED')
    return jsonify({
        'cycle': {'id': active.id, 'name': active.cycle_name, 'status': active.status,
                  'parta_deadline': active.parta_deadline, 'partb_deadline': active.partb_deadline},
        'stats': {
            'total': total,
            'parta_submitted': parta_sub, 'parta_draft': parta_draft, 'parta_not_started': parta_not,
            'parta_pending': total - parta_sub,
            'partb_submitted': partb_sub, 'partb_draft': partb_draft,
            'partb_pending': total - partb_sub,
            'hr_scored': hr_scored, 'hr_pending': total - hr_scored,
        }
    })

# ─── MY FORM ────────────────────────────────────────────────────────────────
@app.route('/api/my-form', methods=['GET'])
@login_required
def my_form():
    active = AppraisalCycle.query.filter_by(status='ACTIVE').first()
    if not active: return jsonify({'error': 'No active appraisal cycle'}), 404
    form = AppraisalForm.query.filter_by(cycle_id=active.id, employee_id=session['employee_id']).first()
    if not form: return jsonify({'error': 'No appraisal form found'}), 404
    ratings = PartARating.query.filter_by(form_id=form.id).all()
    texts = PartAText.query.filter_by(form_id=form.id).all()
    emp = Employee.query.get(session['employee_id'])
    md = EMP_LOOKUP.get(emp.employee_code, {})
    return jsonify({
        'form_id': form.id, 'parta_status': form.parta_status, 'partb_status': form.partb_status,
        'hr_status': form.hr_status, 'final_score': form.final_score, 'final_rating': form.final_rating,
        'question_sections': md.get('a_secs', []),
        'ratings': {f'{r.section_code}_{r.question_index}': r.rating for r in ratings},
        'texts': {f'{t.section_code}_{t.field_key}': t.response_text for t in texts},
        'parta_last_saved': form.parta_last_saved.isoformat() if form.parta_last_saved else None,
        'parta_submitted_at': form.parta_submitted_at.isoformat() if form.parta_submitted_at else None,
        'partb_submitted_at': form.partb_submitted_at.isoformat() if form.partb_submitted_at else None,
    })

@app.route('/api/forms/<int:fid>/parta', methods=['POST'])
@login_required
def save_parta(fid):
    form = AppraisalForm.query.get_or_404(fid)
    if form.employee_id != session['employee_id']: return jsonify({'error': 'Not your form'}), 403
    if form.parta_status == 'SUBMITTED': return jsonify({'error': 'Already submitted'}), 400
    data = request.json or {}
    PartARating.query.filter_by(form_id=fid).delete()
    PartAText.query.filter_by(form_id=fid).delete()
    for key, val in data.get('ratings', {}).items():
        parts = key.split('_')
        if len(parts) >= 2:
            sec, qi = parts[0], int(parts[1])
            qtext = ''
            qs = QUESTION_BANK.get(sec, {}).get('questions', [])
            if 0 < qi <= len(qs): qtext = qs[qi - 1]
            db.session.add(PartARating(form_id=fid, section_code=sec, question_index=qi, question_text=qtext, rating=int(val)))
    for key, val in data.get('texts', {}).items():
        idx = key.find('_')
        if idx > 0:
            db.session.add(PartAText(form_id=fid, section_code=key[:idx], field_key=key[idx+1:], response_text=val))
    now = datetime.datetime.utcnow()
    if data.get('submit'):
        form.parta_status = 'SUBMITTED'
        form.parta_submitted_at = now
    else:
        form.parta_status = 'IN_PROGRESS'
    form.parta_last_saved = now
    db.session.commit()
    return jsonify({'success': True, 'status': form.parta_status, 'last_saved': now.isoformat()})

# ─── TEAM / PART B ──────────────────────────────────────────────────────────
@app.route('/api/team-forms', methods=['GET'])
@login_required
def team_forms():
    active = AppraisalCycle.query.filter_by(status='ACTIVE').first()
    if not active: return jsonify([])
    role = session.get('role')
    emp_id = session['employee_id']
    if role in ('SUPER_ADMIN', 'HR_ADMIN'):
        reports = Employee.query.filter(Employee.is_active == True, Employee.id != emp_id).order_by(Employee.full_name).all()
    else:
        reports = Employee.query.filter(Employee.is_active == True, Employee.reporting_manager_id == emp_id).order_by(Employee.full_name).all()
    result = []
    for e in reports:
        form = AppraisalForm.query.filter_by(cycle_id=active.id, employee_id=e.id).first()
        if not form: continue
        md = EMP_LOOKUP.get(e.employee_code, {})
        result.append({
            'form_id': form.id, 'employee_code': e.employee_code, 'full_name': e.full_name,
            'designation': e.designation, 'vertical': e.department.vertical_code if e.department else '',
            'parta_status': form.parta_status, 'partb_status': form.partb_status, 'hr_status': form.hr_status,
            'question_sections_a': md.get('a_secs', []),
            'question_sections_b': md.get('b_secs', []),
        })
    return jsonify(result)

@app.route('/api/forms/<int:fid>/partb', methods=['GET'])
@login_required
def get_partb(fid):
    form = AppraisalForm.query.get_or_404(fid)
    role = session.get('role')
    emp_id = session['employee_id']
    emp = form.employee
    if role not in ('SUPER_ADMIN', 'HR_ADMIN') and emp.reporting_manager_id != emp_id:
        return jsonify({'error': 'Not authorized'}), 403
    md = EMP_LOOKUP.get(emp.employee_code, {})
    ratings = PartBRating.query.filter_by(form_id=fid).all()
    texts = PartBText.query.filter_by(form_id=fid).all()
    parta_r = PartARating.query.filter_by(form_id=fid).all()
    parta_t = PartAText.query.filter_by(form_id=fid).all()
    return jsonify({
        'form_id': fid, 'employee_code': emp.employee_code, 'full_name': emp.full_name,
        'parta_status': form.parta_status, 'partb_status': form.partb_status,
        'parta_sections': md.get('a_secs', []),
        'partb_sections': md.get('b_secs', []),
        'parta_ratings': {f'{r.section_code}_{r.question_index}': r.rating for r in parta_r},
        'parta_texts': {f'{t.section_code}_{t.field_key}': t.response_text for t in parta_t},
        'partb_ratings': {f'{r.section_code}_{r.question_index}': r.rating for r in ratings},
        'partb_texts': {f'{t.section_code}_{t.field_key}': t.response_text for t in texts},
    })

@app.route('/api/forms/<int:fid>/partb', methods=['POST'])
@login_required
def save_partb(fid):
    form = AppraisalForm.query.get_or_404(fid)
    role = session.get('role')
    emp_id = session['employee_id']
    emp = form.employee
    if role not in ('SUPER_ADMIN', 'HR_ADMIN') and emp.reporting_manager_id != emp_id:
        return jsonify({'error': 'Not authorized'}), 403
    if form.partb_status == 'SUBMITTED': return jsonify({'error': 'Already submitted'}), 400
    data = request.json or {}
    PartBRating.query.filter_by(form_id=fid).delete()
    PartBText.query.filter_by(form_id=fid).delete()
    for key, val in data.get('ratings', {}).items():
        parts = key.split('_')
        if len(parts) >= 2:
            sec, qi = parts[0], int(parts[1])
            qtext = ''
            qs = QUESTION_BANK.get(sec, {}).get('questions', [])
            if 0 < qi <= len(qs): qtext = qs[qi - 1]
            db.session.add(PartBRating(form_id=fid, manager_id=emp_id, section_code=sec, question_index=qi, question_text=qtext, rating=int(val)))
    for key, val in data.get('texts', {}).items():
        idx = key.find('_')
        if idx > 0:
            db.session.add(PartBText(form_id=fid, manager_id=emp_id, section_code=key[:idx], field_key=key[idx+1:], response_text=val))
    now = datetime.datetime.utcnow()
    if data.get('submit'):
        form.partb_status = 'SUBMITTED'
        form.partb_submitted_at = now
    else:
        form.partb_status = 'IN_PROGRESS'
    form.partb_last_saved = now
    db.session.commit()
    return jsonify({'success': True, 'status': form.partb_status, 'last_saved': now.isoformat()})

# ─── QUESTIONS API ──────────────────────────────────────────────────────────
@app.route('/api/questions/<section>', methods=['GET'])
@login_required
def get_questions(section):
    bank = QUESTION_BANK.get(section)
    if not bank: return jsonify({'error': 'Section not found'}), 404
    return jsonify({'section': section, 'title': bank['title'], 'questions': bank['questions'], 'open_text': bank.get('open_text', [])})

# ─── TRACKER & SCORING ──────────────────────────────────────────────────────
@app.route('/api/tracker', methods=['GET'])
@login_required
def tracker():
    if session.get('role') not in ('SUPER_ADMIN', 'HR_ADMIN'): return jsonify({'error': 'Unauthorized'}), 403
    active = AppraisalCycle.query.filter_by(status='ACTIVE').first()
    if not active: return jsonify([])
    forms = AppraisalForm.query.filter_by(cycle_id=active.id).all()
    return jsonify([{
        'form_id': f.id, 'employee_code': f.employee.employee_code, 'full_name': f.employee.full_name,
        'designation': f.employee.designation, 'vertical': f.employee.department.vertical_code if f.employee.department else '',
        'grade': f.employee.grade.grade_code if f.employee.grade else '',
        'manager_name': f.employee.manager.full_name if f.employee.manager else '',
        'parta_status': f.parta_status, 'partb_status': f.partb_status, 'hr_status': f.hr_status,
        'final_score': f.final_score, 'final_rating': f.final_rating,
    } for f in forms])

@app.route('/api/forms/<int:fid>/score', methods=['POST'])
@login_required
def score_form(fid):
    if session.get('role') not in ('SUPER_ADMIN', 'HR_ADMIN'): return jsonify({'error': 'Unauthorized'}), 403
    form = AppraisalForm.query.get_or_404(fid)
    data = request.json or {}
    form.final_score = data.get('final_score')
    form.final_rating = data.get('final_rating')
    form.increment_recommendation = data.get('increment_recommendation')
    form.promotion_recommendation = data.get('promotion_recommendation')
    form.readiness_level = data.get('readiness_level')
    form.hr_notes = data.get('hr_notes')
    form.hr_status = 'SCORED'
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/forms/<int:fid>/detail', methods=['GET'])
@login_required
def form_detail(fid):
    form = AppraisalForm.query.get_or_404(fid)
    emp = form.employee
    md = EMP_LOOKUP.get(emp.employee_code, {})
    return jsonify({
        'form_id': fid, 'employee_code': emp.employee_code, 'full_name': emp.full_name,
        'designation': emp.designation, 'vertical': emp.department.vertical_code if emp.department else '',
        'grade': emp.grade.grade_code if emp.grade else '',
        'manager_name': emp.manager.full_name if emp.manager else '',
        'parta_status': form.parta_status, 'partb_status': form.partb_status, 'hr_status': form.hr_status,
        'final_score': form.final_score, 'final_rating': form.final_rating,
        'increment_recommendation': form.increment_recommendation,
        'promotion_recommendation': form.promotion_recommendation, 'hr_notes': form.hr_notes,
        'question_sections_a': md.get('a_secs', []),
        'question_sections_b': md.get('b_secs', []),
        'parta_ratings': {f'{r.section_code}_{r.question_index}': r.rating for r in PartARating.query.filter_by(form_id=fid).all()},
        'parta_texts': {f'{t.section_code}_{t.field_key}': t.response_text for t in PartAText.query.filter_by(form_id=fid).all()},
        'partb_ratings': {f'{r.section_code}_{r.question_index}': r.rating for r in PartBRating.query.filter_by(form_id=fid).all()},
        'partb_texts': {f'{t.section_code}_{t.field_key}': t.response_text for t in PartBText.query.filter_by(form_id=fid).all()},
    })

# ─── EMPLOYEES & CYCLES ─────────────────────────────────────────────────────
@app.route('/api/employees', methods=['GET'])
@login_required
def get_employees():
    emps = Employee.query.filter_by(is_active=True).order_by(Employee.full_name).all()
    return jsonify([{
        'id': e.id, 'employee_code': e.employee_code, 'full_name': e.full_name,
        'designation': e.designation, 'grade': e.grade.grade_code if e.grade else '',
        'vertical': e.department.vertical_code if e.department else '',
        'manager_name': e.manager.full_name if e.manager else '',
    } for e in emps])

@app.route('/api/admin/update_employee', methods=['POST'])
@login_required
def admin_update_employee():
    if session.get('employee_code') != 'DIR12010':
        return jsonify({'error': 'Only Super Admin (DIR12010) can edit employee data'}), 403
    data = request.json or {}
    emp_id = data.get('employee_id')
    if not emp_id:
        return jsonify({'error': 'employee_id required'}), 400
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({'error': 'Employee not found'}), 404
    if 'full_name' in data and data['full_name'].strip():
        emp.full_name = data['full_name'].strip()
    if 'designation' in data:
        emp.designation = data['designation'].strip() if data['designation'] else emp.designation
    if 'reporting_manager_code' in data:
        mgr_code = data['reporting_manager_code'].strip().upper()
        if mgr_code:
            mgr = Employee.query.filter_by(employee_code=mgr_code, is_active=True).first()
            if not mgr:
                return jsonify({'error': f'Manager {mgr_code} not found'}), 400
            if mgr.id == emp.id:
                return jsonify({'error': 'Cannot set self as manager'}), 400
            emp.reporting_manager_id = mgr.id
    db.session.commit()
    return jsonify({'success': True, 'message': f'Updated {emp.employee_code}'})

@app.route('/api/cycles', methods=['GET'])
@login_required
def get_cycles():
    return jsonify([{'id':c.id,'cycle_name':c.cycle_name,'fy_start':c.fy_start,'fy_end':c.fy_end,'status':c.status,
        'parta_deadline':c.parta_deadline,'partb_deadline':c.partb_deadline} for c in AppraisalCycle.query.order_by(AppraisalCycle.id.desc()).all()])

@app.route('/api/cycles/<int:cid>/activate', methods=['POST'])
@login_required
def activate_cycle(cid):
    if session.get('role') not in ('SUPER_ADMIN','HR_ADMIN'): return jsonify({'error':'Unauthorized'}),403
    AppraisalCycle.query.filter(AppraisalCycle.status=='ACTIVE').update({'status':'CLOSED'})
    cycle = AppraisalCycle.query.get_or_404(cid)
    cycle.status = 'ACTIVE'
    for e in Employee.query.filter_by(is_active=True).all():
        if not AppraisalForm.query.filter_by(cycle_id=cid, employee_id=e.id).first():
            db.session.add(AppraisalForm(cycle_id=cid, employee_id=e.id))
    db.session.commit()
    return jsonify({'success': True})

# ─── EXPORT ──────────────────────────────────────────────────────────────────
@app.route('/api/export', methods=['GET'])
@login_required
def export_excel():
    if session.get('role') not in ('SUPER_ADMIN','HR_ADMIN'): return jsonify({'error':'Unauthorized'}),403
    active = AppraisalCycle.query.filter_by(status='ACTIVE').first()
    wb = Workbook()
    
    navy = PatternFill('solid', fgColor='CC1E2E')
    navy2 = PatternFill('solid', fgColor='1B3A6B')
    green_f = PatternFill('solid', fgColor='E8F5E9')
    amber_f = PatternFill('solid', fgColor='FFF3E0')
    gray_f = PatternFill('solid', fgColor='F5F5F5')
    white_f = PatternFill('solid', fgColor='FFFFFF')
    hf2 = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    bf2 = Font(name='Arial', bold=True, size=9)
    nf2 = Font(name='Arial', size=9)
    sf2 = Font(name='Arial', size=8, color='666666')
    bdr = Border(left=Side('thin',color='CCCCCC'),right=Side('thin',color='CCCCCC'),
                 top=Side('thin',color='CCCCCC'),bottom=Side('thin',color='CCCCCC'))
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ═══ SHEET 1: Summary Tracker ═══
    ws = wb.active; ws.title = 'Tracker Summary'
    headers = ['Emp Code','Full Name','Vertical','Grade','Designation','Manager','Part A','Part B','HR Status','Score','Rating','Increment','Promotion','Notes']
    widths = [13,28,13,7,26,24,13,13,12,8,22,18,18,30]
    for i, (h,w) in enumerate(zip(headers,widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h); c.font = hf2; c.fill = navy; c.alignment = ha; c.border = bdr
    ws.freeze_panes = 'A2'
    forms = AppraisalForm.query.filter_by(cycle_id=active.id).all() if active else []
    for ri, f in enumerate(forms, 2):
        emp = f.employee
        row = [emp.employee_code, emp.full_name, emp.department.vertical_code if emp.department else '',
               emp.grade.grade_code if emp.grade else '', emp.designation or '',
               emp.manager.full_name if emp.manager else '',
               f.parta_status, f.partb_status, f.hr_status, f.final_score, f.final_rating or '',
               f.increment_recommendation or '', f.promotion_recommendation or '', f.hr_notes or '']
        bg = white_f if ri % 2 == 0 else gray_f
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = nf2; c.fill = bg; c.border = bdr
    
    # ═══ SHEET 2: Part A — All Employee Answers ═══
    ws2 = wb.create_sheet('Part A Answers')
    # Build header: Employee info + one column per question
    # First collect all unique section+question combos across all employees
    all_a_questions = []
    seen = set()
    for f in forms:
        md = EMP_LOOKUP.get(f.employee.employee_code, {})
        for sec in md.get('a_secs', []):
            bank = QUESTION_BANK.get(sec, {})
            for qi, q in enumerate(bank.get('questions', []), 1):
                key = f'{sec}_{qi}'
                if key not in seen:
                    seen.add(key)
                    all_a_questions.append((sec, qi, q, key))
            for fk, label in bank.get('open_text', []):
                key = f'{sec}_{fk}'
                if key not in seen:
                    seen.add(key)
                    all_a_questions.append((sec, fk, f'[Text] {label}', key))
    
    # Sort by section then question index
    def sort_key(x):
        sec_num = int(x[0][1:]) if x[0][1:].isdigit() else 99
        qi = x[1] if isinstance(x[1], int) else 999
        return (sec_num, qi)
    all_a_questions.sort(key=sort_key)
    
    # Headers
    info_headers = ['Emp Code', 'Full Name', 'Vertical', 'Grade', 'Manager', 'Status']
    for i, h in enumerate(info_headers, 1):
        c = ws2.cell(row=1, column=i, value=h); c.font = hf2; c.fill = navy; c.alignment = ha; c.border = bdr
        ws2.column_dimensions[get_column_letter(i)].width = [13,26,12,7,22,12][i-1]
    # Question headers
    for qi, (sec, qidx, qtxt, key) in enumerate(all_a_questions):
        col = len(info_headers) + qi + 1
        c = ws2.cell(row=1, column=col, value=f'{sec}-Q{qidx}')
        c.font = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill = navy2; c.alignment = ha; c.border = bdr
        ws2.column_dimensions[get_column_letter(col)].width = 10
        # Question text in row 2
        c2 = ws2.cell(row=2, column=col, value=qtxt[:80])
        c2.font = Font(name='Arial', size=7, color='666666'); c2.alignment = la; c2.border = bdr
    # Info headers in row 2 too
    for i, h in enumerate(info_headers, 1):
        ws2.cell(row=2, column=i, value='').border = bdr
    ws2.freeze_panes = 'G3'
    ws2.row_dimensions[1].height = 24
    ws2.row_dimensions[2].height = 50
    
    # Data rows
    for ri, f in enumerate(forms, 3):
        emp = f.employee; md = EMP_LOOKUP.get(emp.employee_code, {})
        info = [emp.employee_code, emp.full_name, emp.department.vertical_code if emp.department else '',
                emp.grade.grade_code if emp.grade else '',
                emp.manager.full_name if emp.manager else '', f.parta_status]
        bg = white_f if ri % 2 == 0 else gray_f
        for ci, val in enumerate(info, 1):
            c = ws2.cell(row=ri, column=ci, value=val); c.font = nf2; c.fill = bg; c.border = bdr
        
        # Get this employee's answers
        ratings = {f'{r.section_code}_{r.question_index}': r.rating for r in PartARating.query.filter_by(form_id=f.id).all()}
        texts = {f'{t.section_code}_{t.field_key}': t.response_text for t in PartAText.query.filter_by(form_id=f.id).all()}
        
        for qi, (sec, qidx, qtxt, key) in enumerate(all_a_questions):
            col = len(info_headers) + qi + 1
            val = ratings.get(key) or texts.get(key) or ''
            c = ws2.cell(row=ri, column=col, value=val)
            c.font = nf2; c.border = bdr
            if isinstance(val, int):
                c.alignment = ha
                c.fill = green_f if val >= 4 else (amber_f if val >= 3 else gray_f)
            else:
                c.alignment = la; c.fill = bg
    
    # ═══ SHEET 3: Part B — All Manager Answers ═══
    ws3 = wb.create_sheet('Part B Answers')
    all_b_questions = []
    seen_b = set()
    for f in forms:
        md = EMP_LOOKUP.get(f.employee.employee_code, {})
        for sec in md.get('b_secs', []):
            bank = QUESTION_BANK.get(sec, {})
            for qi, q in enumerate(bank.get('questions', []), 1):
                key = f'{sec}_{qi}'
                if key not in seen_b:
                    seen_b.add(key)
                    all_b_questions.append((sec, qi, q, key))
            for fk, label in bank.get('open_text', []):
                key = f'{sec}_{fk}'
                if key not in seen_b:
                    seen_b.add(key)
                    all_b_questions.append((sec, fk, f'[Text] {label}', key))
    all_b_questions.sort(key=sort_key)
    
    for i, h in enumerate(info_headers, 1):
        c = ws3.cell(row=1, column=i, value=h); c.font = hf2; c.fill = navy; c.alignment = ha; c.border = bdr
        ws3.column_dimensions[get_column_letter(i)].width = [13,26,12,7,22,12][i-1]
    for qi, (sec, qidx, qtxt, key) in enumerate(all_b_questions):
        col = len(info_headers) + qi + 1
        c = ws3.cell(row=1, column=col, value=f'{sec}-Q{qidx}')
        c.font = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill = navy2; c.alignment = ha; c.border = bdr
        ws3.column_dimensions[get_column_letter(col)].width = 10
        c2 = ws3.cell(row=2, column=col, value=qtxt[:80])
        c2.font = Font(name='Arial', size=7, color='666666'); c2.alignment = la; c2.border = bdr
    for i in range(1, len(info_headers)+1):
        ws3.cell(row=2, column=i, value='').border = bdr
    ws3.freeze_panes = 'G3'
    ws3.row_dimensions[1].height = 24
    ws3.row_dimensions[2].height = 50
    
    for ri, f in enumerate(forms, 3):
        emp = f.employee; md = EMP_LOOKUP.get(emp.employee_code, {})
        info = [emp.employee_code, emp.full_name, emp.department.vertical_code if emp.department else '',
                emp.grade.grade_code if emp.grade else '',
                emp.manager.full_name if emp.manager else '', f.partb_status]
        bg = white_f if ri % 2 == 0 else gray_f
        for ci, val in enumerate(info, 1):
            c = ws3.cell(row=ri, column=ci, value=val); c.font = nf2; c.fill = bg; c.border = bdr
        ratings = {f'{r.section_code}_{r.question_index}': r.rating for r in PartBRating.query.filter_by(form_id=f.id).all()}
        texts = {f'{t.section_code}_{t.field_key}': t.response_text for t in PartBText.query.filter_by(form_id=f.id).all()}
        for qi, (sec, qidx, qtxt, key) in enumerate(all_b_questions):
            col = len(info_headers) + qi + 1
            val = ratings.get(key) or texts.get(key) or ''
            c = ws3.cell(row=ri, column=col, value=val)
            c.font = nf2; c.border = bdr
            if isinstance(val, int):
                c.alignment = ha
                c.fill = green_f if val >= 4 else (amber_f if val >= 3 else gray_f)
            else:
                c.alignment = la; c.fill = bg
    
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'PRERNA_Full_Export_{datetime.date.today()}.xlsx')

# ─── FULL BACKUP (all Part A + Part B answers) ──────────────────────────
@app.route('/api/backup', methods=['GET'])
@login_required
def backup_data():
    if session.get('role') not in ('SUPER_ADMIN','HR_ADMIN'):
        return jsonify({'error':'Unauthorized'}), 403
    import sqlite3
    db_path = DB_PATH
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    wb2 = Workbook()
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    red_f = PatternFill('solid', fgColor='CC1E2E')
    blue_f = PatternFill('solid', fgColor='1B3A6B')
    grn_f = PatternFill('solid', fgColor='E8F5E9')
    amb_f = PatternFill('solid', fgColor='FFF3E0')
    gry_f = PatternFill('solid', fgColor='F5F5F5')
    wht_f = PatternFill('solid', fgColor='FFFFFF')
    hf = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    nf = Font(name='Arial', size=9)
    bd = Border(left=Side('thin',color='CCCCCC'),right=Side('thin',color='CCCCCC'),
                top=Side('thin',color='CCCCCC'),bottom=Side('thin',color='CCCCCC'))
    
    def write_sheet(ws, query, headers, widths, fill=red_f, rating_col=None):
        for i,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1,column=i,value=h); c.font=hf; c.fill=fill; c.alignment=ha; c.border=bd
        ws.freeze_panes = 'A2'
        cur.execute(query)
        for ri, row in enumerate(cur.fetchall(), 2):
            bg = wht_f if ri%2==0 else gry_f
            for ci, val in enumerate(list(row), 1):
                c = ws.cell(row=ri, column=ci, value=val or '')
                c.font=nf; c.fill=bg; c.border=bd
                if rating_col and ci==rating_col and isinstance(val,int):
                    c.alignment=ha
                    c.fill = grn_f if val>=4 else (amb_f if val==3 else gry_f)
        return ws.max_row - 1
    
    # Sheet 1: Employee Master
    ws1 = wb2.active; ws1.title = 'Employee Master'
    n1 = write_sheet(ws1, """
        SELECT e.employee_code, e.full_name, d.vertical_code, g.grade_code, e.designation,
               m.employee_code, m.full_name, ua.role,
               CASE WHEN ua.password_reset_required=0 THEN 'Yes' ELSE 'No' END,
               ua.last_login
        FROM employees e
        LEFT JOIN departments d ON e.department_id=d.id LEFT JOIN grades g ON e.grade_id=g.id
        LEFT JOIN employees m ON e.reporting_manager_id=m.id LEFT JOIN user_auth ua ON ua.employee_id=e.id
        WHERE e.is_active=1 ORDER BY e.employee_code
    """, ['Emp Code','Full Name','Vertical','Grade','Designation','Mgr Code','Mgr Name','Role','PW Changed','Last Login'],
    [14,28,13,7,26,14,24,14,12,20])
    
    # Sheet 2: Appraisal Status
    ws2 = wb2.create_sheet('Appraisal Status')
    n2 = write_sheet(ws2, """
        SELECT e.employee_code, e.full_name, d.vertical_code, g.grade_code, e.designation,
               m.full_name, af.parta_status, af.parta_submitted_at, af.partb_status, af.partb_submitted_at,
               af.hr_status, af.final_score, af.final_rating, af.increment_recommendation, af.promotion_recommendation, af.hr_notes
        FROM appraisal_forms af JOIN employees e ON af.employee_id=e.id
        JOIN appraisal_cycles ac ON af.cycle_id=ac.id AND ac.status='ACTIVE'
        LEFT JOIN departments d ON e.department_id=d.id LEFT JOIN grades g ON e.grade_id=g.id
        LEFT JOIN employees m ON e.reporting_manager_id=m.id WHERE e.is_active=1 ORDER BY e.employee_code
    """, ['Emp Code','Full Name','Vertical','Grade','Designation','Manager','Part A','Part A Date','Part B','Part B Date','HR','Score','Rating','Increment','Promotion','HR Notes'],
    [14,28,13,7,26,24,13,20,13,20,10,8,22,16,16,30])
    
    # Sheet 3: Part A Ratings
    ws3 = wb2.create_sheet('Part A Ratings')
    n3 = write_sheet(ws3, """
        SELECT e.employee_code, e.full_name, d.vertical_code, par.section_code, par.question_index, par.question_text, par.rating
        FROM part_a_ratings par JOIN appraisal_forms af ON par.form_id=af.id
        JOIN appraisal_cycles ac ON af.cycle_id=ac.id AND ac.status='ACTIVE'
        JOIN employees e ON af.employee_id=e.id LEFT JOIN departments d ON e.department_id=d.id
        ORDER BY e.employee_code, par.section_code, par.question_index
    """, ['Emp Code','Full Name','Vertical','Section','Q#','Question','Rating'],
    [14,28,13,10,5,70,9], rating_col=7)
    
    # Sheet 4: Part A Texts
    ws4 = wb2.create_sheet('Part A Text Responses')
    n4 = write_sheet(ws4, """
        SELECT e.employee_code, e.full_name, d.vertical_code, pat.section_code, pat.field_key, pat.response_text
        FROM part_a_texts pat JOIN appraisal_forms af ON pat.form_id=af.id
        JOIN appraisal_cycles ac ON af.cycle_id=ac.id AND ac.status='ACTIVE'
        JOIN employees e ON af.employee_id=e.id LEFT JOIN departments d ON e.department_id=d.id
        ORDER BY e.employee_code, pat.section_code
    """, ['Emp Code','Full Name','Vertical','Section','Field','Response'],
    [14,28,13,10,24,80])
    
    # Sheet 5: Part B Ratings
    ws5 = wb2.create_sheet('Part B Ratings')
    n5 = write_sheet(ws5, """
        SELECT e.employee_code, e.full_name, d.vertical_code, mgr.employee_code, mgr.full_name,
               pbr.section_code, pbr.question_index, pbr.question_text, pbr.rating
        FROM part_b_ratings pbr JOIN appraisal_forms af ON pbr.form_id=af.id
        JOIN appraisal_cycles ac ON af.cycle_id=ac.id AND ac.status='ACTIVE'
        JOIN employees e ON af.employee_id=e.id LEFT JOIN departments d ON e.department_id=d.id
        LEFT JOIN employees mgr ON pbr.manager_id=mgr.id
        ORDER BY e.employee_code, pbr.section_code, pbr.question_index
    """, ['Emp Code','Full Name','Vertical','Rated By Code','Rated By Name','Section','Q#','Question','Rating'],
    [14,28,13,14,24,10,5,70,9], fill=blue_f, rating_col=9)
    
    # Sheet 6: Part B Texts
    ws6 = wb2.create_sheet('Part B Text Responses')
    n6 = write_sheet(ws6, """
        SELECT e.employee_code, e.full_name, d.vertical_code, mgr.employee_code, mgr.full_name,
               pbt.section_code, pbt.field_key, pbt.response_text
        FROM part_b_texts pbt JOIN appraisal_forms af ON pbt.form_id=af.id
        JOIN appraisal_cycles ac ON af.cycle_id=ac.id AND ac.status='ACTIVE'
        JOIN employees e ON af.employee_id=e.id LEFT JOIN departments d ON e.department_id=d.id
        LEFT JOIN employees mgr ON pbt.manager_id=mgr.id
        ORDER BY e.employee_code, pbt.section_code
    """, ['Emp Code','Full Name','Vertical','Rated By Code','Rated By Name','Section','Field','Response'],
    [14,28,13,14,24,10,24,80], fill=blue_f)
    
    conn.close()
    output = io.BytesIO(); wb2.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'PRERNA_Full_Backup_{datetime.date.today()}.xlsx')

# ─── STATIC & FRONTEND ──────────────────────────────────────────────────
@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

@app.route('/')
@app.route('/<path:path>')
def index(path=''):
    return render_template('index.html')


# ─── SEED DATABASE (clean — no pre-filled data) ─────────────────────────────
def seed_database():
    print("Seeding Procam PMS v5.0 (124 validated employees, 32 question sections)...")
    c1 = Company(name='Procam Logistics Pvt. Ltd.', short_code='PLPL')
    c2 = Company(name='Procam Worldwide Logistics Pvt. Ltd.', short_code='PWLPL')
    db.session.add_all([c1, c2]); db.session.flush()
    loc = Location(company_id=c1.id, city='Kolkata', state='West Bengal', branch_name='Head Office')
    db.session.add(loc); db.session.flush()
    grades_data = [('C2','Consultant',1),('J2','Junior 2',2),('J1','Junior 1',3),('T1','Trainee',4),
                   ('E1','Executive',5),('M3','Manager',6),('M2','Senior Manager',7),('M1','Director/VP',8)]
    grade_map = {}
    for gc, gn, so in grades_data:
        g = Grade(grade_code=gc, grade_name=gn, sort_order=so); db.session.add(g); db.session.flush(); grade_map[gc] = g.id
    verticals = ['Installation','PTM','PTM-M','PFM','Warehouse','Corporate','IT','HR','Finance','Admin']
    dept_map = {}
    for v in verticals:
        cid = c1.id if v in ('Installation','PTM','PTM-M','Warehouse') else c2.id
        d = Department(company_id=cid, dept_name=v, vertical_code=v); db.session.add(d); db.session.flush(); dept_map[v] = d.id
    emp_map = {}
    for code, name, vert, sf, grade, desig, mgr, role, a_secs, b_secs in EMPLOYEE_MASTER:
        dept_id = dept_map.get(sf, dept_map.get(vert, dept_map['Corporate']))
        emp = Employee(employee_code=code, full_name=name, designation=desig,
                       grade_id=grade_map.get(grade, grade_map['E1']),
                       department_id=dept_id, company_id=c1.id, location_id=loc.id,
                       reporting_manager_id=emp_map.get(mgr), is_active=True)
        db.session.add(emp); db.session.flush(); emp_map[code] = emp.id
        ua = UserAuth(employee_id=emp.id, password_hash=bcrypt.generate_password_hash(code).decode(),
                      role=role, password_reset_required=True)
        db.session.add(ua)
    # Fix manager references
    for code, name, vert, sf, grade, desig, mgr, role, a_secs, b_secs in EMPLOYEE_MASTER:
        if mgr and mgr in emp_map:
            emp = Employee.query.filter_by(employee_code=code).first()
            if emp and not emp.reporting_manager_id:
                emp.reporting_manager_id = emp_map[mgr]
    db.session.flush()
    cycle = AppraisalCycle(cycle_name='FY 2025-26', fy_start='2025-04-01', fy_end='2026-03-31',
                           parta_deadline='2026-04-30', partb_deadline='2026-05-15',
                           results_publish_date='2026-06-01', status='ACTIVE')
    db.session.add(cycle); db.session.flush()
    for code, emp_id in emp_map.items():
        db.session.add(AppraisalForm(cycle_id=cycle.id, employee_id=emp_id))
    db.session.commit()
    total = Employee.query.filter_by(is_active=True).count()
    print(f"Seeded: {total} employees | {AppraisalForm.query.count()} forms | FY 2025-26 ACTIVE")
    print("All users must change password on first login. No pre-filled data.")

def init_db():
    db.create_all()
    if Company.query.count() == 0:
        seed_database()

def migrate_db():
    """Safe migration — adds new columns to existing DB without losing data."""
    import sqlite3
    db_path = DB_PATH
    if not os.path.exists(db_path):
        return
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    # Check existing columns in appraisal_forms
    cursor.execute("PRAGMA table_info(appraisal_forms)")
    existing_cols = {row[1] for row in cursor.fetchall()}
    # Add missing columns (non-destructive — nullable, default NULL)
    migrations = [
        ('parta_last_saved', 'DATETIME'),
        ('partb_last_saved', 'DATETIME'),
    ]
    for col_name, col_type in migrations:
        if col_name not in existing_cols:
            try:
                cursor.execute(f"ALTER TABLE appraisal_forms ADD COLUMN {col_name} {col_type}")
                print(f"  Migration: added column appraisal_forms.{col_name}")
            except Exception as e:
                print(f"  Migration skip: {col_name} — {e}")
    conn.commit()
    conn.close()

_db_initialized = False

@app.before_request
def ensure_db():
    global _db_initialized
    if not _db_initialized:
        try: Company.query.count()
        except Exception: init_db()
        _db_initialized = True

with app.app_context():
    migrate_db()  # Safe column additions — runs before init_db
    init_db()     # Creates tables only if DB is empty

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)